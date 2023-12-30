
from flask import Flask, render_template, request, redirect, url_for, jsonify, Response, make_response
import os

from PyPDF2 import PdfFileReader
from docx import Document
from openpyxl import load_workbook
from pptx import Presentation
from PIL import Image
from moviepy.editor import VideoFileClip
from werkzeug.utils import secure_filename
from mutagen.mp3 import MP3
from mutagen.easyid3 import EasyID3
from mutagen.flac import FLAC
from mutagen.wavpack import WavPack
from mutagen.aiff import AIFF
from moviepy.editor import VideoFileClip
from datetime import datetime
from PyPDF2.generic import IndirectObject
from io import StringIO
app = Flask(__name__)
UPLOAD_FOLDER = 'uploads'
app.config['UPLOAD_FOLDER'] = UPLOAD_FOLDER
all_file_metadata = []

@app.route('/')
def index():
    return render_template('index.html')
@app.route('/upload', methods=['POST'])
def upload():
    if 'folder' in request.files:
        uploaded_files = request.files.getlist('folder')
        temp_folder = UPLOAD_FOLDER
        os.makedirs(temp_folder, exist_ok=True)
        for uploaded_file in uploaded_files:
            filename = secure_filename(uploaded_file.filename)
            file_path = os.path.join(temp_folder, filename)
            uploaded_file.save(file_path)

        all_file_metadata.clear()
        for root, dirs, files in os.walk(temp_folder):
            for file_name in files:
                file_path = os.path.join(root, file_name)
                metadata = extract_metadata(file_path)
                if metadata:
                   
                    all_file_metadata.append({'file_name': file_name, 'metadata': metadata})

        return redirect(url_for('metadata', folder_path=temp_folder))
    else:
        return render_template('index.html', error="No folder selected")


@app.route('/metadata/<folder_path>')
def metadata(folder_path):
    return render_template('metadata.html', all_file_metadata=all_file_metadata)
def simple_convert(metadata):
    if isinstance(metadata, dict):
        return {k: str(v) for k, v in metadata.items()}
    elif isinstance(metadata, list):
        return [str(elem) for elem in metadata]
    else:
        return str(metadata)

def simple_convert(metadata):
    if isinstance(metadata, dict):
        return {k: convert_value(v) for k, v in metadata.items()}
    elif isinstance(metadata, list):
        return [convert_value(item) for item in metadata]
    else:
        return convert_value(metadata)

def convert_value(value):
    if isinstance(value, datetime):
        return value.strftime('%Y-%m-%d %H:%M:%S')
    elif isinstance(value, IndirectObject):
        
        return str(value)
    elif isinstance(value, (list, dict)):
        return simple_convert(value)
    elif value is None:
        return 'None'
    else:
        return value
@app.route('/metadata_data')
def metadata_data():
    for item in all_file_metadata:
        print(item)  
    return jsonify(simple_convert(all_file_metadata))
def convert_to_text(metadata):
    lines = []
    for item in metadata:
        lines.append(f"File Name: {item['file_name']}")
        for key, value in item['metadata'].items():
            lines.append(f"{key}: {value}")
        lines.append("\n")  
    return "\n".join(lines)

import csv
from flask import Response

@app.route('/download_metadata')
def download_metadata():
    si = StringIO()
    cw = csv.writer(si)
    
    # En-têtes des colonnes
    cw.writerow(['File Name', 'Attribute', 'Value'])

    for file_data in all_file_metadata:
        for key, value in file_data['metadata'].items():
            cw.writerow([file_data['file_name'], key, value])

    output = make_response(si.getvalue())
    output.headers["Content-Disposition"] = "attachment; filename=metadata.csv"
    output.headers["Content-type"] = "text/csv"
    return output

def extract_docx_metadata(file_path):
    try:
        doc = Document(file_path)
        return {
            'Title': doc.core_properties.title or "Inconnu",
            'Author': doc.core_properties.author or "Inconnu",
            'Created': doc.core_properties.created or "Inconnu",
            'Modified': doc.core_properties.modified or "Inconnu",
            'Subject': doc.core_properties.subject or "Inconnu",
            'Comments': doc.core_properties.comments or "Inconnu",
            'Category': doc.core_properties.category or "Inconnu",
            'Keywords': doc.core_properties.keywords or "Inconnu",
            'Last Modified By': doc.core_properties.last_modified_by or "Inconnu",
            'Revision': doc.core_properties.revision or "Inconnu"
            
        }
    except Exception as e:
        return {"Error": str(e)}



from openpyxl import load_workbook

def extract_xlsx_metadata(file_path):
    try:
        excel_doc = load_workbook(file_path)
        properties = excel_doc.properties
        return {
            'Title': properties.title or "Inconnu",
            'Author': properties.author or "Inconnu",
            'Created': properties.created or "Inconnu",
            'Modified': properties.modified or "Inconnu",
            'Last Modified By': properties.last_modified_by or "Inconnu",
            'Category': properties.category or "Inconnu",
            'Keywords': properties.keywords or "Inconnu",
            'Comments': properties.comments or "Inconnu",
            'Subject': properties.subject or "Inconnu"
            
        }
    except Exception as e:
        return {"Error": str(e)}


def extract_pptx_metadata(file_path):
    try:
        pptx_doc = Presentation(file_path)
        return {
            'Title': pptx_doc.core_properties.title or "Inconnu",
            'Author': pptx_doc.core_properties.author or "Inconnu",
            'Created': pptx_doc.core_properties.created or "Inconnu",
            'Modified': pptx_doc.core_properties.modified or "Inconnu",
            'Last Modified By': pptx_doc.core_properties.last_modified_by or "Inconnu",
            'Revision': pptx_doc.core_properties.revision or "Inconnu",
            'Subject': pptx_doc.core_properties.subject or "Inconnu",
            'Keywords': pptx_doc.core_properties.keywords or "Inconnu",
            'Comments': pptx_doc.core_properties.comments or "Inconnu",
            'Category': pptx_doc.core_properties.category or "Inconnu"
          
        }
    except Exception as e:
        return {"Error": str(e)}

def extract_pdf_metadata(pdf_path):
    with open(pdf_path, 'rb') as pdf_file:
        pdf_reader = PdfFileReader(pdf_file)
        info = pdf_reader.getDocumentInfo()
        return simple_convert(info)        
from PIL import Image
from PIL.ExifTags import TAGS, GPSTAGS

def get_decimal_from_dms(dms, ref):
    degrees, minutes, seconds = dms
    decimal = degrees + (minutes / 60.0) + (seconds / 3600.0)
    if ref in ['S', 'W']:
        decimal = -decimal
    return decimal

def get_gps_coordinates(exif_data):
    if 'GPSInfo' in exif_data:
        gps_info = exif_data['GPSInfo']
        gps_latitude = get_decimal_from_dms(gps_info[2], gps_info[1])
        gps_longitude = get_decimal_from_dms(gps_info[4], gps_info[3])
        return gps_latitude, gps_longitude
    return None, None

def extract_image_metadata(image_path):
    with Image.open(image_path) as img:
        exif_data = {}
        if hasattr(img, '_getexif'):
            exif_info = img._getexif()
            if exif_info is not None:
                for tag, value in exif_info.items():
                    decoded = TAGS.get(tag, tag)
                    exif_data[decoded] = value

        gps_latitude, gps_longitude = get_gps_coordinates(exif_data)

        image_metadata = {
            'Format': img.format,
            'Mode': img.mode,
            'Size': img.size,
            'Info': img.info,
            'GPS Latitude': gps_latitude,
            'GPS Longitude': gps_longitude,
        }
        return image_metadata

    return {"Error": "Unable to open image or read EXIF data"}
def extract_audio_metadata(audio_path):
    file_extension = audio_path.split('.')[-1].lower()
    metadata = {}

    try:
        if file_extension == "mp3":
            audio = MP3(audio_path, ID3=EasyID3)
        elif file_extension == "flac":
            audio = FLAC(audio_path)
        elif file_extension == "wv":
            audio = WavPack(audio_path)
        elif file_extension == "aiff":
            audio = AIFF(audio_path)
        else:
            return {"Error": "Unsupported audio format"}

        for key in audio.keys():
            metadata[key] = audio[key]

        return metadata
    except Exception as e:
        return {"Error": str(e)}



def format_duration(seconds):
    hours = int(seconds // 3600)
    minutes = int((seconds % 3600) // 60)
    seconds = seconds % 60
    return f"{hours}h {minutes}m {seconds:.2f}s"


def extract_video_metadata(video_path):
    with VideoFileClip(video_path) as clip:
        width, height = clip.size
        aspect_ratio = f"{width}:{height}"

        video_metadata = {
            'Duration (s)': clip.duration,
            'Readable Duration': format_duration(clip.duration),
            'Width': width,
            'Height': height,
            'Aspect Ratio': aspect_ratio,
            'FPS': clip.fps,
        }

        if hasattr(clip.reader, 'bitrate'):
            video_metadata['Bitrate (kbps)'] = clip.reader.bitrate / 1000
        else:
            video_metadata['Bitrate (kbps)'] = 'Unknown'

        if clip.audio and hasattr(clip.audio, 'nchannels'):
            video_metadata['Audio Channels'] = clip.audio.nchannels
        else:
            video_metadata['Audio Channels'] = 'None'

        if hasattr(clip.reader, 'codec'):
            video_metadata['Video Codec'] = clip.reader.codec
        else:
            video_metadata['Video Codec'] = 'Unknown'

        if clip.audio and hasattr(clip.audio.reader, 'codec'):
            video_metadata['Audio Codec'] = clip.audio.reader.codec
        else:
            video_metadata['Audio Codec'] = 'Unknown'

    return video_metadata


def extract_metadata(file_path):
    file_extension = file_path.split('.')[-1].lower()
    supported_extensions = ['docx', 'xlsx', 'pptx', 'pdf', 'jpg', 'jpeg', 'png', 'gif', 'mp3', 'wav', 'mp4', 'avi', 'mkv']

    if file_extension in supported_extensions:
        if file_extension == 'docx':
            return extract_docx_metadata(file_path)
        elif file_extension == 'xlsx':
            return extract_xlsx_metadata(file_path)
        elif file_extension == 'pptx':
            return extract_pptx_metadata(file_path)
        elif file_extension == 'pdf':
            return extract_pdf_metadata(file_path)
        elif file_extension in ['jpg', 'jpeg', 'png', 'gif']:
            return extract_image_metadata(file_path)
        elif file_extension in ['mp3', 'wav']:
            return extract_audio_metadata(file_path)
        elif file_extension in ['mp4', 'avi', 'mkv']:
            return extract_video_metadata(file_path)
    else:
        return None
if __name__ == '__main__':
    app.run(debug=True)
